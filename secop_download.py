#!/usr/bin/env python

"""
Resumable Socrata dataset downloader for Colombian procurement data.
Features:
- Downloads dataset in batches (configurable chunk_size).
- Tracks progress in an Excel file: Batch, Status, Length, Error.
- Exports a small sample summary to Excel before downloading.
- Uses retry and timeout logic for robustness.
- Writes each batch to a temporary file and appends atomically to the main newline-delimited JSON output.

Defaults and behavior:
- dataset_id default: jbjy-vk9h (configurable)
- chunk_size default: 50000
- Output files placed in current working directory.
- Requires: pandas, sodapy, openpyxl

Run:
    python secop_download.py

"""

import os
import time
from datetime import datetime
import argparse
import logging
from pathlib import Path
import pandas as pd
from sodapy import Socrata
import openpyxl

start_time = time.perf_counter()

# Default configuration (can be overridden by CLI)
DATASET_ID = "jbjy-vk9h"
CHUNK_SIZE = 50000
TIMEOUT = 60
RETRIES = 5
RETRY_DELAY = 10

# Helper: safe get count with retries
def safe_get_count(client, dataset_id, retries=5, delay=10):
    for attempt in range(retries):
        try:
            res = client.get(dataset_id, select="count(*)")
            return int(res[0]["count"])
        except Exception as e:
            logging.warning("Attempt %d/%d to get count failed: %s", attempt + 1, retries, e)
            time.sleep(delay)
    raise RuntimeError("Failed to get count after retries")

# Helper: safe get records with retries
def safe_get_records(client, dataset_id, limit, offset, retries=5, delay=10):
    for attempt in range(retries):
        try:
            return client.get(dataset_id, limit=limit, offset=offset)
        except Exception as e:
            logging.warning("Attempt %d/%d to get records (offset=%s) failed: %s", attempt + 1, retries, offset, e)
            time.sleep(delay)
    logging.error("Giving up on offset=%s after %d attempts", offset, retries)
    return None

# Helper: create or load progress workbook
def load_or_create_progress(filename, total_batches):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Batch", "Status", "Length", "Error"])
        for b in range(total_batches):
            ws.append([b, "Pending", 0, ""])
        wb.save(filename)
    return wb, ws

# Helper: append batch data to main JSON file (newline-delimited JSON)
def append_batch_to_main(json_filename, batch_records):
    # write to main file in append mode
    with open(json_filename, "a", encoding="utf-8") as f:
        for rec in batch_records:
            # convert record (dict) to JSON string via pandas Series for consistent handling
            f.write(f"{pd.Series(rec).to_json(force_ascii=False)}\\n")

# Main process
def main():
    parser = argparse.ArgumentParser(description="Resumable Socrata dataset downloader")
    parser.add_argument("--dataset", default=DATASET_ID, help="Socrata dataset id (default in file)")
    parser.add_argument("--chunk-size", type=int, default=CHUNK_SIZE, help="Batch size (default 50000)")
    parser.add_argument("--timeout", type=int, default=TIMEOUT, help="HTTP read timeout seconds")
    parser.add_argument("--retries", type=int, default=RETRIES, help="Retries for API calls")
    parser.add_argument("--retry-delay", type=int, default=RETRY_DELAY, help="Delay between retries (s)")
    parser.add_argument("--keep-temp", action="store_true", help="Keep per-batch temp files for debugging")
    parser.add_argument("--out-dir", default=".", help="Output directory")
    parser.add_argument("--app-token", default=None, help="Socrata app token (optional)")
    parser.add_argument("--username", default=None, help="Socrata username (optional)")
    parser.add_argument("--password", default=None, help="Socrata password (optional)")
    parser.add_argument("--no-summary", action="store_true", help="Skip sample summary export")
    args = parser.parse_args()

    # configure logging
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    log_fn = out_dir / f"secop_download_{args.dataset}_{datetime.now().strftime('%m%d%Y')}.log"
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(message)s",
                        handlers=[
                            logging.StreamHandler(),
                            logging.FileHandler(log_fn, encoding="utf-8")
                        ])

    logging.info("Starting resumable download for dataset %s", args.dataset)

    # recreate client with provided credentials/token
    client_local = Socrata("www.datos.gov.co", args.app_token or None, username=args.username or None, password=args.password or None, timeout=args.timeout)

    total_rows = safe_get_count(client_local, args.dataset, retries=args.retries, delay=args.retry_delay)
    total_batches = (total_rows + args.chunk_size - 1) // args.chunk_size
    logging.info("Total rows: %d, Total batches (chunk_size=%d): %d", total_rows, args.chunk_size, total_batches)

    # Filenames per run
    today_str = datetime.now().strftime("%m%d%Y")
    json_filename = out_dir / f"{args.dataset}_{today_str}.json"
    progress_xlsx = out_dir / f"progress_{args.dataset}_{today_str}.xlsx"
    summary_xlsx = out_dir / f"summary_{args.dataset}_{today_str}.xlsx"

    # Summary: fetch first 100 records to inspect schema
    if not args.no_summary:
        sample = safe_get_records(client_local, args.dataset, limit=100, offset=0, retries=args.retries, delay=args.retry_delay)
        if sample:
            df_sample = pd.DataFrame(sample)
            logging.info("Sample rows: %d", df_sample.shape[0])
            logging.info("Columns: %s", list(df_sample.columns))
            logging.info("Dtypes:\n%s", df_sample.dtypes)
            df_sample.to_excel(summary_xlsx, index=False)
            logging.info("Saved summary to %s", summary_xlsx)
        else:
            logging.warning("Could not fetch sample records for summary")

    # ensure progress file exists
    wb, ws = load_or_create_progress(str(progress_xlsx), total_batches)

    # Ensure main file exists
    if not json_filename.exists():
        json_filename.write_text("", encoding="utf-8")

    # Iterate over batches according to progress sheet
    for row in ws.iter_rows(min_row=2, max_row=1 + total_batches, values_only=False):
        batch_num = row[0].value
        status = row[1].value
        if status == "Completed":
            continue
        offset = int(batch_num) * args.chunk_size
        logging.info("Processing batch %s/%s (offset=%s) - Status: %s", batch_num + 1, total_batches, offset, status)

        # fetch records for batch
        batch = safe_get_records(client_local, args.dataset, limit=args.chunk_size, offset=offset, retries=args.retries, delay=args.retry_delay)
        if batch is None:
            # mark error
            row[1].value = "Error"
            row[2].value = 0
            row[3].value = "Failed after retries"
            wb.save(str(progress_xlsx))
            logging.error("Batch %s failed after retries; progress saved to %s", batch_num, progress_xlsx)
            continue

        # write batch to temp file first
        temp_fn = out_dir / f"{args.dataset}_batch_{batch_num}_{today_str}.ndjson"
        with open(temp_fn, "w", encoding="utf-8") as tf:
            for rec in batch:
                tf.write(f"{pd.Series(rec).to_json(force_ascii=False)}\n")

        # append temp to main file
        append_batch_to_main(str(json_filename), batch)

        # update progress
        row[1].value = "Completed"
        row[2].value = len(batch)
        row[3].value = ""
        wb.save(str(progress_xlsx))

        # remove temp file unless keep
        if not args.keep_temp:
            try:
                temp_fn.unlink()
            except Exception:
                pass

        logging.info("Batch %s completed (%s records). Progress saved.", batch_num, len(batch))

    logging.info("All batches processed or attempted. Main file: %s", json_filename)
    end_time = time.perf_counter()
    logging.info("Elapsed time: %.2f seconds", end_time - start_time)


if __name__ == "__main__":
    main()
